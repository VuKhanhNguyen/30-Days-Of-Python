#file handling tức là thao tác với file, bao gồm đọc, ghi, xóa file
# 'r' : read - đọc file, nếu file không tồn tại thì sẽ báo lỗi
# 'w' : write - ghi file, nếu file đã tồn tại thì sẽ xóa nội dung cũ và ghi mới
# 'a' : append - ghi file, nếu file đã tồn tại thì sẽ thêm nội dung mới vào cuối file
# 'x' : exclusive creation - tạo file mới, nếu file đã tồn tại thì sẽ báo lỗi
# 't' : text - đọc ghi file dưới dạng text, mặc định là text
# 'b' : binary - đọc ghi file dưới dạng nhị phân, thường dùng cho file hình ảnh, âm thanh, video
# mặc định là 'r'

file_path = 'K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'
f = open(file_path)
print(f)


#read() : đọc toàn bộ nội dung file và trả về chuỗi
file_path = 'K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'
f = open(file_path)
txt = f.read()
print(type(txt))
print(txt)
txt2 = f.read(10) # đọc 10 ký tự đầu tiên
print(txt2) # không có gì vì đã đọc hết nội dung file rồi
f.close()

file_path = 'K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'
f = open(file_path)
txt = f.read(10)
print(type(txt))
print(txt)
f.close()

#readline() : đọc 1 dòng trong file và trả về chuỗi
file_path = 'K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'
f = open(file_path)
line = f.readline()
print(type(line))
print(line)
f.close()

#readlines() : đọc toàn bộ nội dung file và trả về list các dòng
file_path = 'K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'
f = open(file_path)
lines = f.readlines()
print(type(lines))
print(lines)
f.close()

#splitlines() : tách các dòng trong file và trả về list các dòng
file_path = 'K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'
f = open(file_path)
splitline = f.read().splitlines()
print(type(splitline))
print(splitline)
f.close()

#with open() : tự động đóng file sau khi đọc xong, không cần phải gọi f.close()
with open('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt') as f:
    lines = f.read().splitlines()
    print(type(lines))
    print(lines)
    
#append : thêm nội dung vào cuối file    
with open('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt','a') as f:
    f.write('This text has to be appended at the end')

#write : ghi nội dung vào file, nếu file đã tồn tại thì sẽ xóa nội dung cũ và ghi mới
with open('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt','w') as f:
    f.write('This text has to be written in the file')
    
# delete : xóa file, nếu file không tồn tại thì sẽ báo lỗi
import os
os.remove('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt')

import os
if os.path.exists('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt'):
    os.remove('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/reading_file_example.txt')
else: 
    print('Đéo có file này đâu')


#json to dictionary dùng loads()
import json
person_json = '''{
    "name": "Asabeneh",
    "country": "Finland",
    "city": "Helsinki",
    "skills": ["JavaScrip", "React", "Python"]
}'''
person_dct = json.loads(person_json)
print(type(person_dct))
print(person_dct)
print(person_dct['name'])

#dictionary to json dùng dumps()
import json
person = {
    "name": "Asabeneh",
    "country": "Finland",
    "city": "Helsinki",
    "skills": ["JavaScrip", "React", "Python"]
}
person_json = json.dumps(person, indent=4) #indent=4 là khoảng cách giữa các dòng, nếu không có thì sẽ không có khoảng cách giữa các dòng
                                            # indent gồm 2,4,8                             
print(type(person_json))
print(person_json)

#lưu json
import json

person = {
    "name": "Asabeneh",
    "country": "Finland",
    "city": "Helsinki",
    "skills": ["JavaScrip", "React", "Python"]
}

with open('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/json_example.json') as f:
    json.dump(person, f, ensure_ascii=False, indent=4) #ensure_ascii=False là không chuyển đổi sang unicode, nếu không có thì sẽ chuyển đổi sang unicode
    
#csv
import csv
with open('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/csv_example.csv') as f:
    csv_reader = csv.reader(f, delimiter=',') 
    line_count = 0
    for row in csv_reader:
        if line_count == 0:
            print(f'Column names are :{", ".join(row)}')
            line_count += 1
        else:
            print(
                f'\t{row[0]} is a teachers. He lives in {row[1]}, {row[2]}.')
            line_count += 1
    print(f'Number of lines:  {line_count}')

#xlsx
import openpyxl
excel_book = openpyxl.load_workbook('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/excel_example.xlsx')
print(len(excel_book.sheetnames))
print(excel_book.sheetnames)

#xml
import xml.etree.ElementTree as ET
tree = ET.parse('K:/AllOfPython/pythonLearning/30daysofPython/30-Days-Of-Python_NVK/30-Days-Of-Python/files/xml_example.xml')
root = tree.getroot()
print('Root tag:', root.tag)
print('Attribute:', root.attrib)
for child in root:
    print('field: ', child.tag)



    